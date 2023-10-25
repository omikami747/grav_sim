#!/usr/bin/env python3
import openpyxl

# Define the file paths for your data
file1 = '../test/decimal.txt'
file2 = 'expected_result.txt'
file3 = 'verilog_result.txt'


# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Sheet1"

# Define the headers
headers = ["input", "expected", "from_verilog", 'error %']

# Write the headers to the first row
for col_num, header in enumerate(headers, 1):
    sheet.cell(row=1, column=col_num, value=header)

# Read and write data from the files
for file_num, file_path in enumerate([file1, file2, file3], 1):
    with open(file_path, 'r') as file:
        row_num = 2  # Start from the second row
        for line in file:
            line = line.strip()
            sheet.cell(row=row_num, column=file_num, value=line)
            row_num += 1
workbook.save("output.xlsx")

row_num = 2
for row in sheet.iter_rows(min_row=2, values_only=True):
    column_B_value = float(row[1])
    column_C_value = float(row[2])
    sheet.cell(row=row_num, column=4, value=((column_C_value-column_B_value)/column_B_value)*100)
    row_num+=1
# Save the Excel file
workbook.save("output.xlsx")

